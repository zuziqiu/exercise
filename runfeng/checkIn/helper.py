from tkinter import messagebox
from openpyxl.styles import PatternFill, Alignment
from datetime import datetime, timedelta
from constants import green_fill, red_fill, blue_fill

def loop_fill(target_sheet, source_row, is_header):
    # 找到未被使用的新行号
    new_row = target_sheet.max_row + 1  # 获取当前最大行号，并加1
    # 设置原始数据行做参照
    for source_index, source_cell in enumerate(source_row):
        cell = target_sheet.cell(row=new_row, column=source_index + 1)

        if isinstance(source_cell, str) or is_header:
            # 字符串或者参照行则直接填充
            cell.value = source_cell
        elif isinstance(source_cell, dict):
            # 逐个单元格设置值和背景色
            cell.value = source_cell['total_duration']  # 将时长写入未被使用的行
            cell.fill = source_cell['background']  # 设置背景色

        if is_header:
            cell.fill = blue_fill  # 每人的默认参考数据设置蓝色背景
            cell.alignment = Alignment(wrap_text=True, vertical='center')

def calculate(time_group, criteria):
    try:
        # 转换字符串时间为 datetime 对象
        time_objects = [datetime.strptime(t, "%H:%M") for t in time_group]
        # 计算每两个时间之间的时长
        durations = []
        for i in range(0, len(time_objects) - 1, 2):
            duration = (time_objects[i + 1] - time_objects[i]).total_seconds() / 60  # 转换为分钟
            durations.append(duration)

        # 计算总时长，将浮点数转换为 timedelta（以分钟为单位）
        total_duration = sum((timedelta(minutes=value) for value in durations), timedelta())
        total_minutes = total_duration.total_seconds() / 60
        # 转换为小时和分钟
        hours = total_duration.seconds // 3600
        minutes = (total_duration.seconds % 3600) // 60

        if total_minutes > criteria:
            background = green_fill
        else:
            background = red_fill

        return {
            'background': background,
            'total_duration': f'{hours}时{minutes:02d}分'
        }

    except Exception as e:
        messagebox.showerror("calculate_error", f"calculate_error: {e}")

