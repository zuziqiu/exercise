from openpyxl.styles import PatternFill

# 定义背景
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # 黄色
red_fill = PatternFill(start_color='E23614', end_color='E23614', fill_type='solid')  # 红色
orange_fill = PatternFill(start_color='FF8C00', end_color='FF8C00', fill_type='solid')  # 橙色
blue_fill = PatternFill(start_color='4CAFF0', end_color='4CAFF0', fill_type='solid')  # 蓝色
gray_fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')  # 灰色
green_fill = PatternFill(start_color='CEFFCE', end_color='CEFFCE', fill_type='solid')  # 绿色

# 缺卡状态灰色
missing_status = {
    'background': gray_fill,
    'total_duration': None
}
# 非标准次数打卡状态黄色
ignore_status = {
    'background': yellow_fill,
    'total_duration': None
}