import openpyxl
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox
from departurement import distribute
from helper import loop_fill

def select_file():
    try:
        # 打开文件对话框，选择 Excel 文件
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            # 使用 openpyxl 读取 Excel 文件
            wb = load_workbook(file_path)
            source_sheet = wb.active

            # 创建新的工作簿和工作表
            target_wb = openpyxl.Workbook()
            target_sheet = target_wb.active
            target_sheet.column_dimensions['C'].hidden = True

            # 处理数据
            handle_data(source_sheet, target_sheet)

            # 导出到新的 Excel 文件
            output_file = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                         filetypes=[("Excel files", "*.xlsx")])
            if output_file:
                target_wb.save(output_file)
                messagebox.showinfo("成功", f"数据已成功导出到 {output_file}")

    except Exception as e:
        messagebox.showerror("打开文件", f"打开文件发生错误: {e}")

def handle_data(source_sheet, target_sheet):
    try:
        # 复制数据
        for index, source_row in enumerate(source_sheet.iter_rows(values_only=True)):
            if index >= 3:
                handle_per_row(source_row, target_sheet)
            else:
                target_sheet.append(source_row)

    except Exception as e:
        messagebox.showerror("handle_data_error", f"handle_data_error: {e}")

def handle_per_row(source_row, target_sheet):
    try:

        row_time_group, time_count = handle_per_column(source_row)
        time_result = distribute(row_time_group)
        temp_row = []
        if time_result is not None:
            # 复制填充原始数据做参考
            loop_fill(target_sheet, source_row, True)
            for time_index in range(time_count):
                for column_index, column_item in enumerate(row_time_group):
                    if column_index < 3:
                        # 前 3 列的人名、部门、id
                        temp_row.append(source_row[column_index])
                    elif column_item is not None and time_index < len(column_item):
                        temp_row.append((column_item[time_index]))
                    else:
                        temp_row.append(None)

                target_sheet.append(temp_row)
                temp_row = []

            # 追加总时长 item
            loop_fill(target_sheet, time_result, False)

    except Exception as e:
        messagebox.showerror("handle_per_row_error", f"handle_per_row_error: {e}")

def handle_per_column(source_row):
    try:
        row_with_time_array = []
        time_count = 0
        for column_index, source_column in enumerate(source_row):
            if source_column is None:
                row_with_time_array.append(None)
            else:
                # str.strip() 方法去掉切割后每个字符串的前后空格, str.split() 方法根据空格切割字符串
                time_array = [s.strip() for s in source_column.split('\n')]
                time_count = max(time_count, len(time_array))
                row_with_time_array.append(time_array)

        return row_with_time_array, time_count

    except Exception as e:
        messagebox.showerror("handle_per_column_error", f"handle_per_column_error: {e}")



# 创建主窗口
root = tk.Tk()
root.title("Excel 数据导出工具")

# 创建按钮，选择文件
select_button = tk.Button(root, text="选择 Excel 文件并导出", command=select_file)
select_button.pack(pady=20)

# 启动 GUI
root.mainloop()